"""Persistence and queries for tutor-recorded offline marks."""

from __future__ import annotations

import csv
import io
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from statistics import mean

from sqlalchemy.orm import Session

from app.models.content import Batch
from app.models.marks import MarksEntry
from app.models.user import StudentProfile, User


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def pct(scored: float, max_marks: int) -> int:
    if max_marks <= 0:
        return 0
    return round((scored / max_marks) * 100)


def list_marks_entries(
    db: Session,
    institution_id: str,
    *,
    batch_id: str | None = None,
    limit: int = 500,
) -> list[dict]:
    q = db.query(MarksEntry).filter(MarksEntry.institution_id == institution_id)
    if batch_id:
        q = q.filter(MarksEntry.batch_id == batch_id)
    rows = q.order_by(MarksEntry.saved_at.desc()).limit(limit).all()
    return [_entry_dict(db, row) for row in rows]


def list_marks_sessions(
    db: Session,
    institution_id: str,
    *,
    batch_id: str | None = None,
    limit: int = 50,
) -> list[dict]:
    entries = list_marks_entries(db, institution_id, batch_id=batch_id, limit=1000)
    sessions: dict[str, dict] = {}
    for entry in entries:
        sid = entry["sessionId"]
        if sid not in sessions:
            sessions[sid] = {
                "sessionId": sid,
                "assessmentTitle": entry["assessmentTitle"],
                "description": entry.get("description"),
                "batch": entry["batch"],
                "savedAt": entry["savedAt"],
                "source": entry["source"],
                "entries": [],
            }
        sessions[sid]["entries"].append(entry)
    ordered = sorted(sessions.values(), key=lambda s: s["savedAt"], reverse=True)
    return ordered[:limit]


def _entry_dict(db: Session, row: MarksEntry) -> dict:
    student = db.get(StudentProfile, row.student_id)
    name = student.user.name if student and student.user else "Student"
    return {
        "id": row.id,
        "sessionId": row.session_id,
        "studentId": row.student_id,
        "studentName": name,
        "batch": row.batch_name,
        "batchId": row.batch_id,
        "assessmentTitle": row.assessment_title,
        "description": row.description,
        "subject": row.subject,
        "maxMarks": row.max_marks,
        "scoredMarks": row.scored_marks,
        "percentage": row.percentage,
        "source": row.source,
        "conductedOn": row.conducted_on,
        "savedAt": row.saved_at,
    }


def save_marks_bulk(
    db: Session,
    institution_id: str,
    *,
    batch_id: str,
    assessment_title: str,
    description: str | None,
    source: str,
    created_by_user_id: str | None,
    columns: list[dict],
    marks: dict[str, dict[str, str | float]],
    student_ids: list[str],
) -> dict:
    """Persist one save session (multi-column spreadsheet or CSV upload).

    columns: [{ id, subject, conductedOn, maxMarks }]
    marks: { columnId: { studentId: value } }
    """
    batch = db.get(Batch, batch_id)
    if not batch or batch.institution_id != institution_id:
        raise ValueError("Batch not found")

    profiles = {
        p.id: p
        for p in db.query(StudentProfile)
        .join(User)
        .filter(User.institution_id == institution_id, StudentProfile.id.in_(student_ids))
        .all()
    }
    if not profiles:
        raise ValueError("No valid students in batch")

    session_id = f"sess-{uuid.uuid4().hex[:12]}"
    saved_at = _now_iso()
    created: list[MarksEntry] = []

    for col in columns:
        subject = (col.get("subject") or "").strip()
        conducted_on = col.get("conductedOn") or col.get("conducted_on") or saved_at[:10]
        max_marks = int(col.get("maxMarks") or col.get("max_marks") or 0)
        col_id = col.get("id") or col.get("columnId") or subject
        if not subject or max_marks <= 0:
            continue
        col_marks = marks.get(col_id) or {}
        for student_id in student_ids:
            raw = col_marks.get(student_id)
            if raw is None or raw == "":
                continue
            scored = float(raw)
            profile = profiles.get(student_id)
            if not profile:
                continue
            entry = MarksEntry(
                id=f"me-{uuid.uuid4().hex[:10]}",
                institution_id=institution_id,
                session_id=session_id,
                student_id=student_id,
                batch_id=batch_id,
                batch_name=batch.name,
                assessment_title=assessment_title.strip(),
                description=(description or "").strip() or None,
                subject=subject,
                max_marks=max_marks,
                scored_marks=scored,
                percentage=pct(scored, max_marks),
                source=source,
                conducted_on=str(conducted_on)[:16],
                saved_at=saved_at,
                created_by_user_id=created_by_user_id,
            )
            db.add(entry)
            created.append(entry)

    if not created:
        raise ValueError("No marks to save")

    # Refresh batch average from all marks + keep batch avg_score updated
    _refresh_batch_avg_score(db, batch_id, institution_id)
    from app.services.analytics_recompute import recompute_students

    recompute_students(db, institution_id, student_ids, commit=False)
    db.commit()
    return {
        "sessionId": session_id,
        "savedAt": saved_at,
        "count": len(created),
        "entries": [_entry_dict(db, row) for row in created],
    }


def _refresh_batch_avg_score(db: Session, batch_id: str, institution_id: str) -> None:
    batch = db.get(Batch, batch_id)
    if not batch:
        return
    rows = (
        db.query(MarksEntry)
        .filter(MarksEntry.institution_id == institution_id, MarksEntry.batch_id == batch_id)
        .all()
    )
    if rows:
        batch.avg_score = round(mean(r.percentage for r in rows))
        return
    # fallback: leave existing avg_score


def marks_for_students(
    db: Session,
    institution_id: str,
    student_ids: list[str],
    *,
    batch_id: str | None = None,
) -> list[MarksEntry]:
    q = db.query(MarksEntry).filter(
        MarksEntry.institution_id == institution_id,
        MarksEntry.student_id.in_(student_ids),
    )
    if batch_id:
        q = q.filter(MarksEntry.batch_id == batch_id)
    return q.order_by(MarksEntry.conducted_on.asc()).all()


def subject_mastery_from_marks(
    db: Session,
    institution_id: str,
    student_ids: list[str],
    *,
    batch_id: str | None = None,
) -> dict[str, dict[str, list[int]]]:
    """student_id -> subject -> list of percentage scores."""
    result: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for row in marks_for_students(db, institution_id, student_ids, batch_id=batch_id):
        result[row.student_id][row.subject].append(row.percentage)
    return result


def _column_key(entry: dict) -> tuple[str, str, int]:
    return (entry["subject"], str(entry["conductedOn"])[:10], int(entry["maxMarks"]))


def _spreadsheet_from_entries(
    entries: list[dict],
) -> tuple[list[tuple[str, str]], list[dict], dict[tuple[str, tuple[str, str, int]], float]]:
    """Return (students, columns, scores) for wide spreadsheet export."""
    student_map: dict[str, str] = {}
    column_map: dict[tuple[str, str, int], dict] = {}
    scores: dict[tuple[str, tuple[str, str, int]], float] = {}

    for entry in entries:
        student_map[entry["studentId"]] = entry["studentName"]
        key = _column_key(entry)
        if key not in column_map:
            column_map[key] = {
                "subject": entry["subject"],
                "conductedOn": str(entry["conductedOn"])[:10],
                "maxMarks": int(entry["maxMarks"]),
            }
        scores[(entry["studentId"], key)] = float(entry["scoredMarks"])

    students = sorted(student_map.items(), key=lambda item: item[1].lower())
    columns = sorted(
        column_map.values(),
        key=lambda col: (col["subject"].lower(), col["conductedOn"]),
    )
    return students, columns, scores


def _slugify(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in value.strip().lower())[:48] or "export"


def _safe_sheet_title(value: str, used: set[str]) -> str:
    base = "".join(ch for ch in value if ch not in "[]:*?/\\")[:28] or "Marks"
    title = base
    n = 2
    while title in used:
        suffix = f" {n}"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(title)
    return title


def _write_spreadsheet_sheet(ws, entries: list[dict], *, session_meta: dict | None = None) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    students, columns, scores = _spreadsheet_from_entries(entries)
    header_fill = PatternFill("solid", fgColor="FEF3D6")
    header_font = Font(bold=True, color="163A66")
    name_font = Font(bold=True, color="163A66")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    row_idx = 1
    if session_meta:
        ws.cell(row=row_idx, column=1, value=session_meta.get("assessmentTitle") or "Marks")
        ws.cell(row=row_idx, column=2, value=session_meta.get("batch") or "")
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(2, len(columns) + 2))
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="163A66")
        row_idx += 1

    # Row: # | Student | Subject names
    ws.cell(row=row_idx, column=1, value="#")
    ws.cell(row=row_idx, column=2, value="Student")
    for col_i, col in enumerate(columns, start=3):
        cell = ws.cell(row=row_idx, column=col_i, value=col["subject"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    row_idx += 1

    # Row: max marks (/ 50)
    ws.cell(row=row_idx, column=1, value="")
    ws.cell(row=row_idx, column=2, value="")
    for col_i, col in enumerate(columns, start=3):
        cell = ws.cell(row=row_idx, column=col_i, value=f"/ {col['maxMarks']}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    row_idx += 1

    # Row: conducted date
    ws.cell(row=row_idx, column=1, value="")
    ws.cell(row=row_idx, column=2, value="")
    for col_i, col in enumerate(columns, start=3):
        cell = ws.cell(row=row_idx, column=col_i, value=col["conductedOn"])
        cell.fill = header_fill
        cell.font = Font(color="6E8499")
        cell.alignment = center
    row_idx += 1

    ws.cell(row=row_idx - 3, column=1).fill = header_fill
    ws.cell(row=row_idx - 3, column=1).font = header_font
    ws.cell(row=row_idx - 3, column=1).alignment = center
    ws.cell(row=row_idx - 3, column=2).fill = header_fill
    ws.cell(row=row_idx - 3, column=2).font = header_font
    ws.cell(row=row_idx - 3, column=2).alignment = left
    ws.cell(row=row_idx - 2, column=1).fill = header_fill
    ws.cell(row=row_idx - 2, column=2).fill = header_fill
    ws.cell(row=row_idx - 1, column=1).fill = header_fill
    ws.cell(row=row_idx - 1, column=2).fill = header_fill

    for idx, (student_id, name) in enumerate(students, start=1):
        ws.cell(row=row_idx, column=1, value=idx).alignment = center
        name_cell = ws.cell(row=row_idx, column=2, value=name)
        name_cell.font = name_font
        name_cell.alignment = left
        for col_i, col in enumerate(columns, start=3):
            key = (col["subject"], col["conductedOn"], col["maxMarks"])
            scored = scores.get((student_id, key))
            mark_cell = ws.cell(row=row_idx, column=col_i, value=scored if scored is not None else "")
            mark_cell.alignment = center
        row_idx += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for col_i in range(3, len(columns) + 3):
        ws.column_dimensions[get_column_letter(col_i)].width = 16


def export_marks_xlsx(
    db: Session,
    institution_id: str,
    *,
    batch_id: str | None = None,
    session_id: str | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook

    if session_id:
        entries = list_marks_entries(db, institution_id, batch_id=batch_id, limit=5000)
        entries = [e for e in entries if e["sessionId"] == session_id]
        if not entries:
            raise ValueError("No saved marks to export.")
        sessions = [
            {
                "sessionId": session_id,
                "assessmentTitle": entries[0]["assessmentTitle"],
                "batch": entries[0]["batch"],
                "entries": entries,
            }
        ]
    else:
        sessions = list_marks_sessions(db, institution_id, batch_id=batch_id, limit=500)
        if not sessions:
            raise ValueError("No saved marks to export.")

    wb = Workbook()
    used_titles: set[str] = set()
    written = 0
    for session in sessions:
        if not session["entries"]:
            continue
        title = _safe_sheet_title(session["assessmentTitle"], used_titles)
        if written == 0:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        _write_spreadsheet_sheet(ws, session["entries"], session_meta=session)
        written += 1

    if written == 0:
        raise ValueError("No saved marks to export.")

    buffer = io.BytesIO()
    wb.save(buffer)

    if session_id and sessions:
        slug = _slugify(sessions[0]["assessmentTitle"])
    elif batch_id:
        batch = db.get(Batch, batch_id)
        slug = _slugify(batch.name if batch else "batch")
    else:
        slug = "all-marks"

    return buffer.getvalue(), f"prism-marks-{slug}.xlsx"


def export_marks_csv(
    db: Session,
    institution_id: str,
    *,
    batch_id: str | None = None,
    session_id: str | None = None,
) -> tuple[str, str]:
    """Spreadsheet-layout CSV (opens in Excel with the same grid as the UI)."""
    if session_id:
        entries = list_marks_entries(db, institution_id, batch_id=batch_id, limit=5000)
        entries = [e for e in entries if e["sessionId"] == session_id]
        session_groups = [{"entries": entries, "assessmentTitle": entries[0]["assessmentTitle"] if entries else ""}]
    else:
        session_groups = list_marks_sessions(db, institution_id, batch_id=batch_id, limit=500)

    if not session_groups or not any(s.get("entries") for s in session_groups):
        raise ValueError("No saved marks to export.")

    output = io.StringIO()
    writer = csv.writer(output)

    for group_index, session in enumerate(session_groups):
        entries = session.get("entries") or []
        if not entries:
            continue
        if group_index > 0:
            writer.writerow([])
            writer.writerow([])
        if len(session_groups) > 1:
            writer.writerow([session.get("assessmentTitle") or "Marks", session.get("batch") or ""])

        students, columns, scores = _spreadsheet_from_entries(entries)
        writer.writerow(["#", "Student", *[col["subject"] for col in columns]])
        writer.writerow(["", "", *[f"/ {col['maxMarks']}" for col in columns]])
        writer.writerow(["", "", *[col["conductedOn"] for col in columns]])
        for idx, (student_id, name) in enumerate(students, start=1):
            row = [idx, name]
            for col in columns:
                key = (col["subject"], col["conductedOn"], col["maxMarks"])
                scored = scores.get((student_id, key))
                row.append(scored if scored is not None else "")
            writer.writerow(row)

    slug = "export"
    if session_id and session_groups[0].get("entries"):
        slug = _slugify(session_groups[0]["entries"][0]["assessmentTitle"])
    elif batch_id:
        batch = db.get(Batch, batch_id)
        if batch:
            slug = _slugify(batch.name)
    return output.getvalue(), f"prism-marks-{slug}.csv"
